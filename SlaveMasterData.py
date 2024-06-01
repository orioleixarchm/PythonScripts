#Loading required packages
import pandas as pd
import openpyxl as opx
from openpyxl.styles import PatternFill
from datetime import date

#Loading data
with pd.ExcelFile("C:/Users/IU3309/PycharmProjects/pythonProject/Data/MASTER_SLAVE _ ALL_ACR.xlsx") as xls:
    df=pd.read_excel(xls, sheet_name="Data", engine="openpyxl")
    assets=pd.read_excel(xls, sheet_name="Asset Data", engine="openpyxl")

#Cleaning data
df['CREATED ON']=pd.to_datetime(df['CREATED ON'], format='%d.%m.%Y').dt.date
df=df[(df['ORIGIN'].isin(['DN','CN','NN','RN','UN'])) & (df['CREATED ON']>=date(2023,11,1))].reset_index(drop=True)
df=df.dropna(subset=['ASSET NAME']).reset_index(drop=True)
df=df.dropna(subset=['MACROPOL']).reset_index(drop=True)
df=df[(df['MACROPOL']!=999) & (df['MACROPOL']!='999')]
df=df.fillna('')
df=df.sort_values(by=["CREATED ON",'ANSWER']).reset_index()

#Defining PCODE column
AssetTypes=['CONN','EQ','P','PU','PUB','PUC','PUT','PUTM','SPAN','SPANEL','SPANG','SUP','TELECOM']
for index, row in df.iterrows():
    if ' LK' in row['ASSET NAME']:
        df.at[index, 'PCODE']='C'
    elif ' LI ' in row['ASSET NAME']:
        df.at[index, 'PCODE']='L'
    elif row['ASSET NAME'] in ['', ' ']:
        df.at[index, 'PCODE']=''
    else:
        df.at[index, 'PCODE'] = 'P'

print(f'\n Data with PCODE Defined:\n\n\n, {df.head(10)}\n\n with {df.shape[0]} rows and {df.shape[1]} columns.\n')

#Creating grouping Conditions
def grouping(row):
    if row['PCODE'] in ['C', 'L']:
        circuit=row['ASSET NAME'].split()[2]
        return str(row['PCODE']) + "|" + str(row['U']) + "|" + str(circuit) + "|" + str(row['MACROPOL'])
    elif row['PCODE']=='P':
        return str(row['PCODE']) + "|" + str(row['P']) + "|" + str(row['U']) + "|" + str(row['T']) + "|" + str(row['M']) + "|" + str(row['MACROPOL'])
    else:
        return ''

df['Condition']=df.apply(grouping, axis=1)

#Counting Masters per grouping
dfMaster=df[df['ANSWER']=='A']
print(f'\nNumber of Masters ACR per grouping:\n\n\n{df[df["ANSWER"]=="A"]["Condition"].value_counts().head(10)}')
Masters=dfMaster.groupby('Condition')['ACR ID'].first().reset_index(drop=True).to_list()
First=df.groupby('Condition')['ACR ID'].first().reset_index(drop=True).to_list()
contains_slave=df[df['ANSWER']=='S']['Condition'].unique().tolist()
print(f'\nThe following groups contain Slaves: \n')
[print(x) for x in contains_slave]

#Checking for single element groups and eliminate status
groups=df['Condition'].value_counts()
single_groups=groups[groups==1].index.tolist()
print(f'\nNumber of single element groups: {len(single_groups)}')
del_status=['ACRCLS','ACRDEL','ACRFRM']

#Check creation date per group
groups_distinct_dates=df.groupby('Condition')['CREATED ON'].nunique()
print(f'\n Groups with more than one date: {len(groups_distinct_dates[groups_distinct_dates>1])} \n\n {groups_distinct_dates[groups_distinct_dates>1]}.\n')

#Reassingning masters
for index, row in df.iterrows():
    if row['ACR ID'] in Masters:
        df.at[index, 'To Slave'] = 'No'
    elif row['ACR ID'] not in Masters and row['ANSWER']=='S':
        df.at[index, 'To Slave'] = 'Already Slave'
    else:
        df.at[index, 'To Slave'] = 'Yes'

#Exclusion
df=df[~df['Condition'].isin(single_groups)]
df=df[~df['STAT'].isin(del_status)]

#Preparing export data
df=df.sort_values(by=['Condition','CREATED ON','ANSWER'])
new_order=['ACR ID','ORIGIN','ACR AM ID','NEED TYPE','ANSWER','ASSET NAME','MACROPOL','P','U','T','M',
           'PCODE','STAT','Condition','To Slave','CREATED ON','GUID','REQUESTOR','POLICY','INTRF GUID','ID']
df=df[new_order].reset_index(drop=True)

print(f'\n Data Exported:\n, {df.head(10)}\n\n with {df.shape[0]} rows and {df.shape[1]} columns.\n\n\n')

#Exporting to excel
path='C:/Users/IU3309/PycharmProjects/pythonProject/Data/New_Masters.xlsx'
df.to_excel(path, index=False)


#Applying format
wb=opx.load_workbook(path)
ws=wb.active
fill_row=PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
for index, row in df.iterrows():
    if row['ACR ID'] in First:
        for cell in ws[index+2]:
            cell.fill=fill_row

wb.save(path)